from collections import Counter
from openpyxl import load_workbook

# Load the workbook and select the sheet
wb = load_workbook('INTER_106320750.xlsx')
ws = wb['inter_106320750']

# Specify the necessary columns
Geocodigo = 'F'  # Column F (Geocodigo)
Area_Int_column = 'M'  # Column M (Area_Int)
Enq_Legal_column = 'L'  # Column L (Enq_Legal)
Area_column = 'H'  # Column H (Area)
area_cor_column = 'E'  # Column E (Area_COR)
a_int_os_column = 'N'  # Column N (A_INT_OS)
par_num_os_column = 'Q'  # Column Q (Par_Num_OS)

# Dictionary to store distinct values, rows, sum, difference, old and new values, and difference_left
distinct_values = {}

# Iterate through the rows in the Excel file
for row in ws.iter_rows(min_row=2):
    row_num = row[0].row
    geocodigo = ws[f'{Geocodigo}{row_num}'].value
    area_int = ws[f'{Area_Int_column}{row_num}'].value
    enq_legal = ws[f'{Enq_Legal_column}{row_num}'].value
    area_value = ws[f'{Area_column}{row_num}'].value  # Get Area value (H)
    area_cor_value = ws[f'{area_cor_column}{row_num}'].value
    a_int_os_value = ws[f'{a_int_os_column}{row_num}'].value

    # Initialize values if they are None
    area_int = area_int if area_int is not None else 0

    # Skip null (None) values in the distinct column
    if geocodigo is None:
        continue

    # Check if Enq_Legal is greater than Area and use Area instead if necessary
    if enq_legal > area_value:
        enq_legal = area_value

    # If the distinct value already exists, append the row and continue summing column Area_Int_column
    if geocodigo in distinct_values:
        distinct_values[geocodigo]['rows'].append(row_num)
        distinct_values[geocodigo]['sum_1'] += area_int
    else:
        # Initialize the entry with the row number, sum from column Area_Int_column, and the single value from Enq_Legal_column
        distinct_values[geocodigo] = {
            'rows': [row_num],
            'sum_1': area_int,  # Sum of values from column Area_Int_column
            'enq_legal': enq_legal,  # Use either Enq_Legal or Area if Enq_Legal is higher
            'old_A_INT_OS': [],  # Store old A_INT_OS values
            'new_A_INT_OS': [],  # Store new A_INT_OS values
            'difference_left': 0  # Initialize remaining difference
        }

    # Ensure that the old value of A_INT_OS is captured before the update
    distinct_values[geocodigo]['old_A_INT_OS'].append(a_int_os_value)

# Calculate the difference for each distinct geocodigo value
for geocodigo, data in distinct_values.items():
    data['difference'] = round(data['enq_legal'] - data['sum_1'], 4)

# Create a new variable to hold row-level data preserving the Excel row order
rows_with_difference = []

# Iterate through the rows again to build the new data structure
for row in ws.iter_rows(min_row=2):
    row_num = row[0].row
    geocodigo = ws[f'{Geocodigo}{row_num}'].value
    par_num_os = ws[f'{par_num_os_column}{row_num}'].value
    area_int = ws[f'{Area_Int_column}{row_num}'].value
    enq_legal = ws[f'{Enq_Legal_column}{row_num}'].value
    area_value = ws[f'{Area_column}{row_num}'].value  # Get Area value (H)
    a_int_os_value = ws[f'{a_int_os_column}{row_num}'].value
    area_cor_value = ws[f'{area_cor_column}{row_num}'].value
    # Skip null (None) values in the distinct column
    if geocodigo is None:
        continue
    # Get the difference value from distinct_values for this geocodigo
    difference = distinct_values[geocodigo]['difference']

    # Add row data to the new list
    rows_with_difference.append({
        'row_num': row_num,
        'geocodigo': geocodigo,
        'par_num_os': par_num_os,
        'area_int': area_int,
        'enq_legal': enq_legal if enq_legal <= area_value else area_value,  # Ensure Enq_Legal is adjusted
        'a_int_os': a_int_os_value,
        'area_cor': area_cor_value,
        'difference': difference  # Assign the calculated difference value
    })

# Count occurrences of each par_num_os
par_num_os_counts = Counter(row['par_num_os'] for row in rows_with_difference)

# Sort rows based on the count of par_num_os (those with fewer rows come first)
sorted_rows = sorted(rows_with_difference, key=lambda x: par_num_os_counts[x['par_num_os']])

# Now, you can print or use the sorted_rows list
for row_data in sorted_rows:
    print(row_data)
