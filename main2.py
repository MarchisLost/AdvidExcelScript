from collections import Counter
from openpyxl import load_workbook

# Load the workbook and select the sheet
wb = load_workbook('INTER_106320750.xlsx')
ws = wb['inter_106320750']

# Specify the necessary columns
Geocodigo = 'F'  # Column F (Geocodigo)
Area_Int_column = 'M'  # Column M (Area_Int)
Enq_Legal_column = 'L'  # Column L (Enq_Legal)
Area_column = 'H'  # Column H (Area)
area_cor_column = 'E'  # Column E (Area_COR)
a_int_os_column = 'N'  # Column N (A_INT_OS)
par_num_os_column = 'Q'  # Column Q (Par_Num_OS)
d_aio_ac_column = 'O'  # Column O (D_aio_ac)

# Dictionary to store distinct values, rows, sum, difference, old and new values, and difference_left
distinct_values = {}

# Iterate through the rows in the Excel file
for row in ws.iter_rows(min_row=2):
    row_num = row[0].row
    geocodigo = ws[f'{Geocodigo}{row_num}'].value
    area_int = ws[f'{Area_Int_column}{row_num}'].value
    enq_legal = ws[f'{Enq_Legal_column}{row_num}'].value
    area = ws[f'{Area_column}{row_num}'].value
    area_cor = ws[f'{area_cor_column}{row_num}'].value
    a_int_os = ws[f'{a_int_os_column}{row_num}'].value
    d_aio_ac_value = ws[f'{d_aio_ac_column}{row_num}'].value

    # Initialize values if they are None
    area_int = area_int if area_int is not None else 0
    d_aio_ac_value = abs(d_aio_ac_value) if d_aio_ac_value is not None else 0

    # Skip null (None) values in the distinct column
    if geocodigo is None:
        continue

    # Check if Enq_Legal is greater than Area and use Area instead if necessary
    if enq_legal > area:
        enq_legal = area

    # If the distinct value already exists, append the row and continue summing column Area_Int_column
    if geocodigo in distinct_values:
        distinct_values[geocodigo]['rows'].append(row_num)
        distinct_values[geocodigo]['sum_1'] += area_int
    else:
        # Initialize the entry with the row number, sum from column Area_Int_column, and the single value from Enq_Legal_column
        distinct_values[geocodigo] = {
            'rows': [row_num],
            'sum_1': area_int,  # Sum of values from column Area_Int_column
            'enq_legal': enq_legal,  # Store Enq_Legal
            'old_A_INT_OS': [],  # Store old A_INT_OS values
            'n_A_INT_OS': [],  # Store new A_INT_OS values
            'Diff': 0,
            'Diff_left': 0,  # Initialize remaining difference
            'd_aio_ac': [],  # Store D_aio_ac values
            'area_cor': area_cor  # Store Area_COR value
        }

    # Ensure that the old value of A_INT_OS is captured before the update
    distinct_values[geocodigo]['old_A_INT_OS'].append(a_int_os)

# Calculate the difference for each distinct geocodigo value
for geocodigo, data in distinct_values.items():
    data['Diff'] = round(data['enq_legal'] - data['sum_1'], 4)

# Create a new variable to hold row-level data preserving the Excel row order
dict_final_results = []

# Iterate through the rows again to build the new data structure
for row in ws.iter_rows(min_row=2):
    row_num = row[0].row
    geocodigo = ws[f'{Geocodigo}{row_num}'].value
    par_num_os = ws[f'{par_num_os_column}{row_num}'].value
    area_int = ws[f'{Area_Int_column}{row_num}'].value
    enq_legal = ws[f'{Enq_Legal_column}{row_num}'].value
    area = ws[f'{Area_column}{row_num}'].value  # Get Area value (H)
    a_int_os = ws[f'{a_int_os_column}{row_num}'].value
    area_cor = ws[f'{area_cor_column}{row_num}'].value
    d_aio_ac = ws[f'{d_aio_ac_column}{row_num}'].value

    # Initialize values if they are None
    area_int = area_int if area_int is not None else 0
    d_aio_ac = abs(d_aio_ac) if d_aio_ac is not None else 0

    # Skip null (None) values in the distinct column
    if geocodigo is None:
        continue

    # Get the difference value from distinct_values for this geocodigo
    difference = distinct_values[geocodigo]['Diff']

    # Add row data to the new list
    dict_final_results.append({
        'Geo': geocodigo,
        'par_num_os': par_num_os,
        'A_int': area_int,
        'N_A_int': area_int,
        'E_leg': enq_legal,
        'a_int_os': a_int_os,
        'A_cor': area_cor,
        'd_aio_ac': d_aio_ac,
        'Diff': difference,
        'R_Diff': difference
    })

# Count occurrences of each par_num_os
par_num_os_counts = Counter(row['par_num_os'] for row in dict_final_results)

# Sort rows based on the count of par_num_os (those with fewer rows come first)
sorted_rows = sorted(dict_final_results, key=lambda x: par_num_os_counts[x['par_num_os']])

# Now, you can print or use the sorted_rows list
for row_data in sorted_rows:
    # TODO Check if the sum of all area_int of that par_num_os is less or the same of the area_cor
    # TODO Remove the diff for all of the same geocodigo
    if row_data["Diff"] < row_data["d_aio_ac"]:
        row_data["N_A_int"] = round(row_data["N_A_int"] + row_data["R_Diff"], 4)
        row_data["R_Diff"] = 0
    else:
        row_data["N_A_int"] = round(row_data["N_A_int"] + row_data["d_aio_ac"], 4)
        row_data["R_Diff"] = round(row_data["R_Diff"] - row_data["d_aio_ac"], 4)

    # Update R_Diff for all rows with the same geocodigo
    for other_row in sorted_rows:
        if other_row['Geo'] == row_data['Geo']:
            other_row['R_Diff'] = row_data["R_Diff"]

    # Print all the columns
    print(f"Geo: {row_data['Geo']}, Par_num_os: {row_data['par_num_os']}, A_int: {row_data['A_int']}, N_A_int: {row_data['N_A_int']}, E_leg: {row_data['E_leg']}, a_int_os: {row_data['a_int_os']}, A_cor: {row_data['A_cor']}, d_aio_ac: {row_data['d_aio_ac']}, Diff: {row_data['Diff']}, R_Diff: {row_data['R_Diff']}")
